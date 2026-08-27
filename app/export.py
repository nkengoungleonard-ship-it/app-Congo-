import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import os

BLEU_FONCE = "0D3D5C"
ORANGE = "F5A623"
DOSSIER_APP = os.path.dirname(os.path.abspath(__file__))
CHEMIN_LOGO = os.path.join(DOSSIER_APP, 'static', 'images', 'logo.jpg')


def generer_excel_ventes(ventes):
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventes"

    entetes = ["Date", "Référence", "Commercial", "Niveau", "Produit", "Quantité", "Montant (FCFA)", "Taux (%)", "Commission (FCFA)"]
    ws.append(entetes)

    for col in range(1, len(entetes) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=BLEU_FONCE, end_color=BLEU_FONCE, fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for v in ventes:
        ws.append([
            v.date_vente.strftime('%d/%m/%Y %H:%M'),
            v.commercial.reference,
            f"{v.commercial.prenom} {v.commercial.nom}",
            v.commercial.niveau,
            v.produit.nom,
            v.quantite,
            round(v.montant, 0),
            round(v.taux_applique * 100, 0),
            round(v.commission_calculee, 0)
        ])

    for col_cells in ws.columns:
        longueur_max = max(len(str(c.value)) if c.value else 0 for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = longueur_max + 4

    fichier = io.BytesIO()
    wb.save(fichier)
    fichier.seek(0)
    return fichier


def generer_excel_commerciaux(commerciaux):
    wb = Workbook()
    ws = wb.active
    ws.title = "Commerciaux"

    entetes = ["Référence", "Nom", "Prénom", "Niveau", "Ville", "Zone", "Téléphone", "Statut",
               "CA Jour (FCFA)", "CA Mois (FCFA)", "CA Année (FCFA)"]
    ws.append(entetes)

    for col in range(1, len(entetes) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=BLEU_FONCE, end_color=BLEU_FONCE, fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for c in commerciaux:
        ws.append([
            c.reference,
            c.nom,
            c.prenom,
            c.niveau,
            c.ville or "",
            c.zone or "",
            c.telephone or "",
            "Actif" if c.actif else "Inactif",
            round(c.chiffre_affaire_jour(), 0),
            round(c.chiffre_affaire_mois(), 0),
            round(c.chiffre_affaire_annee(), 0)
        ])

    for col_cells in ws.columns:
        longueur_max = max(len(str(c.value)) if c.value else 0 for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = longueur_max + 4

    fichier = io.BytesIO()
    wb.save(fichier)
    fichier.seek(0)
    return fichier


def _entete_pdf(titre):
    styles = getSampleStyleSheet()
    elements = []

    if os.path.exists(CHEMIN_LOGO):
        logo = Image(CHEMIN_LOGO, width=2*cm, height=2*cm)
        logo.hAlign = 'CENTER'
        elements.append(logo)
        elements.append(Spacer(1, 0.2*cm))

    style_titre = ParagraphStyle('titre', parent=styles['Title'], fontSize=16,
                                  textColor=colors.HexColor(f"#{BLEU_FONCE}"), alignment=TA_CENTER)
    elements.append(Paragraph("WaterLife Group International", style_titre))
    elements.append(Paragraph(titre, ParagraphStyle('sous', parent=styles['Normal'],
                                                      fontSize=11, textColor=colors.grey,
                                                      alignment=TA_CENTER, spaceAfter=15)))
    return elements


def generer_pdf_ventes(ventes):
    fichier = io.BytesIO()
    doc = SimpleDocTemplate(fichier, pagesize=landscape(A4), topMargin=1.5*cm, bottomMargin=1.5*cm)

    elements = _entete_pdf(f"Historique des ventes — généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}")

    data = [["Date", "Référence", "Commercial", "Produit", "Qté", "Montant", "Taux", "Commission"]]
    for v in ventes:
        data.append([
            v.date_vente.strftime('%d/%m/%Y'),
            v.commercial.reference,
            f"{v.commercial.prenom} {v.commercial.nom}",
            v.produit.nom,
            str(v.quantite),
            f"{v.montant:.0f}",
            f"{v.taux_applique*100:.0f}%",
            f"{v.commission_calculee:.0f}"
        ])
    if len(data) == 1:
        data.append(["-", "-", "Aucune vente enregistrée", "-", "-", "-", "-", "-"])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(f"#{BLEU_FONCE}")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#dddddd")),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f6f9")]),
        ('ALIGN', (4, 0), (-1, -1), 'CENTER'),
    ]))
    elements.append(table)
    doc.build(elements)
    fichier.seek(0)
    return fichier


def generer_pdf_commerciaux(commerciaux):
    fichier = io.BytesIO()
    doc = SimpleDocTemplate(fichier, pagesize=landscape(A4), topMargin=1.5*cm, bottomMargin=1.5*cm)

    elements = _entete_pdf(f"Liste des commerciaux — généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}")

    data = [["Référence", "Nom complet", "Niveau", "Ville/Zone", "Téléphone", "Statut", "CA Mois"]]
    for c in commerciaux:
        data.append([
            c.reference,
            f"{c.prenom} {c.nom}",
            c.niveau,
            f"{c.ville or ''} {('- ' + c.zone) if c.zone else ''}",
            c.telephone or "-",
            "Actif" if c.actif else "Inactif",
            f"{c.chiffre_affaire_mois():.0f} FCFA"
        ])
    if len(data) == 1:
        data.append(["-", "Aucun commercial enregistré", "-", "-", "-", "-", "-"])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(f"#{BLEU_FONCE}")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#dddddd")),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f6f9")]),
    ]))
    elements.append(table)
    doc.build(elements)
    fichier.seek(0)
    return fichier


def generer_pdf_fiche_commercial(commercial, ventes):
    """Fiche PDF individuelle : infos du commercial + historique de ses ventes.
    Le logo reste centre en haut (position fixe), la photo (si presente) est
    placee dans le coin superieur gauche, independamment du logo."""
    fichier = io.BytesIO()
    doc = SimpleDocTemplate(fichier, pagesize=A4, topMargin=1.5*cm, bottomMargin=1.5*cm,
                             leftMargin=1.8*cm, rightMargin=1.8*cm)
    styles = getSampleStyleSheet()
    elements = []

    chemin_photo = None
    if commercial.photo_filename:
        chemin_photo_test = os.path.join(DOSSIER_APP, 'static', 'uploads', 'commerciaux', commercial.photo_filename)
        if os.path.exists(chemin_photo_test):
            chemin_photo = chemin_photo_test

    def dessiner_entete(canvas, doc_ref):
        largeur_page, hauteur_page = A4
        taille_logo = 2.5 * cm
        x_logo = (largeur_page - taille_logo) / 2
        y_images = hauteur_page - doc_ref.topMargin - taille_logo

        if os.path.exists(CHEMIN_LOGO):
            canvas.drawImage(CHEMIN_LOGO, x_logo, y_images, width=taille_logo, height=taille_logo,
                              preserveAspectRatio=True, mask='auto')

        if chemin_photo:
            taille_photo = 3.5 * cm
            x_photo = doc_ref.leftMargin
            canvas.drawImage(chemin_photo, x_photo, y_images, width=taille_photo, height=taille_photo,
                              preserveAspectRatio=True, mask='auto')

    # Espace reserve en haut de page pour le logo/la photo dessines par dessiner_entete()
    elements.append(Spacer(1, 3.3 * cm))

    style_titre = ParagraphStyle('titre', parent=styles['Title'], fontSize=16,
                                  textColor=colors.HexColor(f"#{BLEU_FONCE}"), alignment=TA_CENTER)
    elements.append(Paragraph("WaterLife Group International", style_titre))
    elements.append(Paragraph("Fiche Commercial — Congo Brazzaville",
                               ParagraphStyle('sous', parent=styles['Normal'], fontSize=11,
                                              textColor=colors.grey, alignment=TA_CENTER, spaceAfter=15)))

    style_ref = ParagraphStyle('ref', parent=styles['Normal'], fontSize=20, alignment=TA_CENTER,
                                textColor=colors.HexColor(f"#{ORANGE}"), fontName='Helvetica-Bold', spaceAfter=15)
    elements.append(Paragraph(commercial.reference, style_ref))

    infos = [
        ["Nom complet", f"{commercial.prenom} {commercial.nom}"],
        ["Niveau", commercial.niveau],
        ["Ville / Zone", f"{commercial.ville or '-'} {('- ' + commercial.zone) if commercial.zone else ''}"],
        ["Téléphone", commercial.telephone or "-"],
        ["Statut", "Actif" if commercial.actif else "Inactif"],
        ["Date de recrutement", commercial.date_recrutement.strftime('%d/%m/%Y')],
        ["Frais d'inscription", "Payés (6 000 FCFA)" if commercial.frais_inscription_paye else "Non payés"],
    ]
    if commercial.parrain:
        infos.append(["Parrainé par", f"{commercial.parrain.prenom} {commercial.parrain.nom} ({commercial.parrain.reference})"])
    if commercial.coordinateur:
        infos.append(["Rattaché au coordinateur", f"{commercial.coordinateur.prenom} {commercial.coordinateur.nom}"])

    table_infos = Table(infos, colWidths=[6*cm, 10*cm])
    table_infos.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.HexColor(f"#{BLEU_FONCE}")),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('LINEBELOW', (0, 0), (-1, -1), 0.5, colors.HexColor("#dddddd")),
    ]))
    elements.append(table_infos)
    elements.append(Spacer(1, 0.6*cm))

    chiffres = [
        ["CA du jour", f"{commercial.chiffre_affaire_jour():.0f} FCFA"],
        ["CA du mois", f"{commercial.chiffre_affaire_mois():.0f} FCFA"],
        ["CA de l'année", f"{commercial.chiffre_affaire_annee():.0f} FCFA"],
    ]
    table_chiffres = Table(chiffres, colWidths=[8*cm, 8*cm])
    table_chiffres.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor("#f4f6f9")),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor(f"#{BLEU_FONCE}")),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor("#dddddd")),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#dddddd")),
    ]))
    elements.append(table_chiffres)
    elements.append(Spacer(1, 0.8*cm))

    progression = commercial.progression_paliers()
    if progression:
        elements.append(Paragraph("Progression vers les paliers de commission", ParagraphStyle(
            'h2b', parent=styles['Heading2'], fontSize=13, textColor=colors.HexColor(f"#{BLEU_FONCE}"), spaceAfter=10)))

        data_progression = [["Produit", "Unités ce mois", "Taux actuel", "Reste pour palier suivant"]]
        for p in progression:
            if p['taux_suivant']:
                reste = f"{p['unites_restantes']} unité(s) pour {p['taux_suivant']*100:.0f}%"
            else:
                reste = "Palier maximum atteint"
            data_progression.append([
                p['produit'],
                str(p['unites']),
                f"{p['taux_actuel']*100:.0f}%",
                reste
            ])

        table_progression = Table(data_progression, colWidths=[5*cm, 3.5*cm, 3*cm, 4.5*cm])
        table_progression.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(f"#{ORANGE}")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#dddddd")),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f6f9")]),
        ]))
        elements.append(table_progression)
        elements.append(Spacer(1, 0.8*cm))

    elements.append(Paragraph("Historique des ventes", ParagraphStyle(
        'h2', parent=styles['Heading2'], fontSize=13, textColor=colors.HexColor(f"#{BLEU_FONCE}"), spaceAfter=10)))

    data_ventes = [["Date", "Produit", "Qté", "Montant", "Taux", "Commission"]]
    for v in ventes:
        data_ventes.append([
            v.date_vente.strftime('%d/%m/%Y'),
            v.produit.nom,
            str(v.quantite),
            f"{v.montant:.0f} FCFA",
            f"{v.taux_applique*100:.0f}%",
            f"{v.commission_calculee:.0f} FCFA"
        ])
    if len(data_ventes) == 1:
        data_ventes.append(["-", "Aucune vente enregistrée", "-", "-", "-", "-"])

    table_ventes = Table(data_ventes, repeatRows=1)
    table_ventes.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(f"#{BLEU_FONCE}")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#dddddd")),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f6f9")]),
    ]))
    elements.append(table_ventes)

    doc.build(elements, onFirstPage=dessiner_entete, onLaterPages=dessiner_entete)
    fichier.seek(0)
    return fichier

def generer_excel_stock(donnees, periode_label):
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock"

    ws.append([f"État du stock — {periode_label}"])
    ws.merge_cells('A1:E1')
    ws['A1'].font = Font(bold=True, size=13, color=BLEU_FONCE)
    ws.append([])

    entetes = ["N°", "Produit", "Stock initial", "Sorties", "Stock actuel"]
    ws.append(entetes)

    ligne_entete = ws.max_row
    for col in range(1, len(entetes) + 1):
        cell = ws.cell(row=ligne_entete, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=BLEU_FONCE, end_color=BLEU_FONCE, fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for i, d in enumerate(donnees, start=1):
        ws.append([
            i,
            d['nom'],
            d['stock_initial'],
            d['sorties'],
            d['stock_actuel']
        ])

    largeurs = [8, 30, 14, 12, 14]
    for idx, largeur in enumerate(largeurs, start=1):
        lettre = ws.cell(row=1, column=idx).column_letter if idx == 1 else chr(64 + idx)
        ws.column_dimensions[chr(64 + idx)].width = largeur

    fichier = io.BytesIO()
    wb.save(fichier)
    fichier.seek(0)
    return fichier

def generer_pdf_stock(donnees, periode_label):
    fichier = io.BytesIO()
    doc = SimpleDocTemplate(fichier, pagesize=landscape(A4), topMargin=1.5*cm, bottomMargin=1.5*cm)

    elements = _entete_pdf(f"État du stock — {periode_label}")

    data = [["N°", "Produit", "Stock initial", "Sorties", "Stock actuel"]]
    for i, d in enumerate(donnees, start=1):
        data.append([
            str(i),
            d['nom'],
            str(d['stock_initial']),
            str(d['sorties']),
            str(d['stock_actuel'])
        ])
    if len(data) == 1:
        data.append(["-", "Aucun produit enregistré", "-", "-", "-"])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(f"#{BLEU_FONCE}")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#dddddd")),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f6f9")]),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
    ]))
    elements.append(table)
    doc.build(elements)
    fichier.seek(0)
    return fichier

def generer_excel_mouvements(mouvements, periode_label):
    wb = Workbook()
    ws = wb.active
    ws.title = "Mouvements"

    ws.append([f"Journal des mouvements de stock — {periode_label}"])
    ws.merge_cells('A1:E1')
    ws['A1'].font = Font(bold=True, size=13, color=BLEU_FONCE)
    ws.append([])

    entetes = ["Date", "Produit", "Type", "Quantité", "Motif", "Enregistré par"]
    ws.append(entetes)

    ligne_entete = ws.max_row
    for col in range(1, len(entetes) + 1):
        cell = ws.cell(row=ligne_entete, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=BLEU_FONCE, end_color=BLEU_FONCE, fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for m in mouvements:
        ws.append([
            m.date_mouvement.strftime('%d/%m/%Y %H:%M'),
            m.produit.nom,
            "Entrée" if m.type_mouvement == 'Entree' else "Sortie",
            m.quantite,
            m.motif or "",
            m.enregistre_par or ""
        ])

    largeurs = [18, 30, 12, 12, 30, 22]
    for idx, largeur in enumerate(largeurs, start=1):
        ws.column_dimensions[chr(64 + idx)].width = largeur

    fichier = io.BytesIO()
    wb.save(fichier)
    fichier.seek(0)
    return fichier


def generer_pdf_mouvements(mouvements, periode_label):
    fichier = io.BytesIO()
    doc = SimpleDocTemplate(fichier, pagesize=landscape(A4), topMargin=1.5*cm, bottomMargin=1.5*cm)

    elements = _entete_pdf(f"Journal des mouvements de stock — {periode_label}")

    data = [["Date", "Produit", "Type", "Quantité", "Motif", "Enregistré par"]]
    for m in mouvements:
        data.append([
            m.date_mouvement.strftime('%d/%m/%Y %H:%M'),
            m.produit.nom,
            "Entrée" if m.type_mouvement == 'Entree' else "Sortie",
            str(m.quantite),
            m.motif or "-",
            m.enregistre_par or "-"
        ])
    if len(data) == 1:
        data.append(["-", "Aucun mouvement enregistré", "-", "-", "-", "-"])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(f"#{BLEU_FONCE}")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#dddddd")),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f6f9")]),
    ]))
    elements.append(table)
    doc.build(elements)
    fichier.seek(0)
    return fichier